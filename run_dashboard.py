#!/usr/bin/env python3
"""
SHSM Analytics Dashboard — Data Builder & Server
=================================================
Reads SEC NoCAs from SEC_NoCA_Dashboard.xlsm, writes data/sec_nocas.json,
then starts a local HTTP server and opens the dashboard in your browser.

Usage:
    python run_dashboard.py            # Refresh data + open hub
    python run_dashboard.py --data     # Refresh data only (no server)
    python run_dashboard.py --serve    # Serve only (skip data refresh)

Requirements:
    pip install openpyxl
"""

import json, re, sys, os, webbrowser, threading, time
sys.stdout.reconfigure(encoding='utf-8')
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run:  pip install openpyxl")
    sys.exit(1)

# ─── Config ───────────────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).parent
EXCEL_FILE   = SCRIPT_DIR / "SEC_NoCA_Dashboard.xlsm"
OUTPUT_JSON  = SCRIPT_DIR / "data" / "sec_nocas.json"
SERVER_PORT  = 8080
OPEN_PAGE    = "index.html"

# ─── Excel reader ──────────────────────────────────────────────────────────────

def fmt_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip() or None

def read_excel(path: Path) -> list[dict]:
    """Read the Data sheet from SEC_NoCA_Dashboard.xlsm and return records."""
    print(f"  Reading {path.name}...")
    wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=False)

    if 'Data' not in wb.sheetnames:
        raise ValueError(f"No 'Data' sheet found in {path.name}. Sheets: {wb.sheetnames}")

    ws = wb['Data']
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        notice_no = row[0]
        if not notice_no:
            continue

        # Clean notice number (strip trailing "New" label)
        notice_no = re.sub(r'\s+New\s*$', '', str(notice_no).strip(), flags=re.IGNORECASE).strip()
        if not re.match(r'^\d{4}-\d{3,4}', notice_no):
            continue

        notice_date = row[7]  # Column H — Notice Date
        if not isinstance(notice_date, datetime):
            continue

        records.append({
            'noticeNo':   notice_no,
            'action':     str(row[1])[:300] if row[1] else '',
            'url':        str(row[2])       if row[2] else '',
            'caseNumber': str(row[3])[:200] if row[3] else '',
            'court':      str(row[4])[:200] if row[4] else '',
            'dateFiled':  fmt_date(row[5]),
            'dateOrder':  fmt_date(row[6]),
            'date':       notice_date.strftime('%Y-%m-%d'),
            'year':       notice_date.year,
            'month':      notice_date.month,
            'claimDue':   fmt_date(row[8]),
        })

    wb.close()

    # Deduplicate by noticeNo, keeping first occurrence
    seen = set()
    deduped = []
    for r in records:
        if r['noticeNo'] not in seen:
            seen.add(r['noticeNo'])
            deduped.append(r)

    deduped.sort(key=lambda r: r['date'])
    return deduped

# ─── Data writer ──────────────────────────────────────────────────────────────

def write_json(records: list[dict]):
    OUTPUT_JSON.parent.mkdir(exist_ok=True)
    payload = {
        'source':       EXCEL_FILE.name,
        'generatedAt':  datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'totalRecords': len(records),
        'records':      records,
    }
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))
    size_kb = OUTPUT_JSON.stat().st_size // 1024
    print(f"  Saved {len(records)} records to {OUTPUT_JSON.relative_to(SCRIPT_DIR)}  ({size_kb} KB)")

# ─── Local HTTP server ─────────────────────────────────────────────────────────

def start_server(port: int):
    """Start a simple HTTP server in SCRIPT_DIR on the given port."""
    import http.server
    os.chdir(SCRIPT_DIR)

    class QuietHandler(http.server.SimpleHTTPRequestHandler):
        def log_message(self, fmt, *args):
            pass  # suppress per-request logging

    server = http.server.HTTPServer(('', port), QuietHandler)
    print(f"  Server running at http://localhost:{port}/")
    print(f"  Press Ctrl+C to stop.\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    data_only  = '--data'  in args
    serve_only = '--serve' in args

    print("=" * 55)
    print("  SHSM Analytics Dashboard")
    print("=" * 55)
    print()

    # Step 1 — Refresh data from Excel
    if not serve_only:
        if not EXCEL_FILE.exists():
            print(f"ERROR: {EXCEL_FILE.name} not found in {SCRIPT_DIR}")
            sys.exit(1)

        print("Reading data from Excel...")
        try:
            records = read_excel(EXCEL_FILE)
        except Exception as e:
            print(f"ERROR reading Excel: {e}")
            sys.exit(1)

        print(f"  {len(records)} NoCAs  ({records[0]['date']} → {records[-1]['date']})")
        write_json(records)
        print()

        if data_only:
            print("Done. Run without --data to also start the server.")
            return

    # Step 2 — Start HTTP server + open browser
    url = f"http://localhost:{SERVER_PORT}/{OPEN_PAGE}"
    print(f"Opening {url} ...")

    # Open browser after a short delay to let the server start
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    start_server(SERVER_PORT)

if __name__ == "__main__":
    main()
